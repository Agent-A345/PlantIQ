import os, sys, time, hashlib, logging, argparse, json, pickle
from pathlib import Path
from dataclasses import dataclass, field
import PyPDF2
from docx import Document as DocxDocument
import openpyxl
import faiss
import numpy as np

try:
    import pytesseract
    from PIL import Image
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)-8s  %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("ingest")

@dataclass
class IngestConfig:
    input_dir: str = "./docs"
    db_dir: str = "./voyage_faiss_db"
    chunk_size: int = 500
    chunk_overlap: int = 50
    min_chunk_words: int = 30
    embedding_backend: str = "voyage"
    plant_id: str = "default"  # multi-tenant: which plant/site these docs belong to
    supported_extensions: list = field(default_factory=lambda: [".pdf",".docx",".txt",".md",".xlsx"])

@dataclass
class DocumentChunk:
    chunk_id: str
    text: str
    source_file: str
    doc_type: str
    page_or_sheet: str
    chunk_index: int
    total_chunks: int
    file_hash: str
    word_count: int
    plant_id: str = "default"  # multi-tenant: which plant/site this chunk belongs to

DIM_HASH = 384
DIM_VOYAGE = 1024
DIM_SENTENCE = 384

def get_dim(backend):
    if backend == "voyage": return DIM_VOYAGE
    if backend == "sentence": return DIM_SENTENCE
    return DIM_HASH

def extract_pdf(path):
    pages = []
    try:
        reader = PyPDF2.PdfReader(str(path))
        for i, page in enumerate(reader.pages):
            pages.append(((page.extract_text() or "").strip(), f"page {i+1}"))
    except Exception as e:
        log.error(f"PDF failed {path.name}: {e}")
    return pages

def extract_docx(path):
    try:
        doc = DocxDocument(str(path))
        blocks, current, idx = [], [], 1
        for para in doc.paragraphs:
            s = para.text.strip()
            if not s: continue
            current.append(s)
            if len(current) >= 30:
                blocks.append(("\n".join(current), f"section {idx}"))
                current, idx = [], idx+1
        if current: blocks.append(("\n".join(current), f"section {idx}"))
        return blocks
    except Exception as e:
        log.error(f"DOCX failed {path.name}: {e}")
        return []

def extract_xlsx(path):
    results = []
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        for name in wb.sheetnames:
            rows = [" | ".join(str(c) for c in row if c is not None and str(c).strip())
                    for row in wb[name].iter_rows(values_only=True)]
            rows = [r for r in rows if r]
            if rows: results.append(("\n".join(rows), name))
        wb.close()
    except Exception as e:
        log.error(f"XLSX failed {path.name}: {e}")
    return results

def extract_text(path):
    try:
        return [(path.read_text(encoding="utf-8", errors="replace").strip(), "full document")]
    except Exception as e:
        log.error(f"TXT failed {path.name}: {e}")
        return []

def extract_document(path, cfg):
    ext = path.suffix.lower()
    if ext == ".pdf": return extract_pdf(path)
    if ext == ".docx": return extract_docx(path)
    if ext == ".xlsx": return extract_xlsx(path)
    if ext in (".txt", ".md"): return extract_text(path)
    return []

def chunk_text(text, chunk_size, overlap, min_words):
    words = text.split()
    if not words: return []
    chunks, start = [], 0
    while start < len(words):
        end = min(start + chunk_size, len(words))
        w = words[start:end]
        if len(w) >= min_words: chunks.append(" ".join(w))
        if end >= len(words): break
        start += chunk_size - overlap
    return chunks

def embed_hash(texts):
    import struct
    vecs = []
    for text in texts:
        seed = text.encode("utf-8")
        vec = [float(struct.unpack_from("<i", hashlib.sha256(seed + i.to_bytes(4,"little")).digest()[:4])[0]) / (2**31) for i in range(DIM_HASH)]
        arr = np.array(vec, dtype=np.float32)
        norm = np.linalg.norm(arr)
        vecs.append(arr / norm if norm > 0 else arr)
    return np.stack(vecs)

def load_voyage_keys():
    """Load all available Voyage API keys — up to 9 different accounts."""
    keys = []
    for var in ["VOYAGE_API_KEY", "VOYAGE_API_KEY_2", "VOYAGE_API_KEY_3",
                "VOYAGE_API_KEY_4", "VOYAGE_API_KEY_5", "VOYAGE_API_KEY_6", "VOYAGE_API_KEY_7", "VOYAGE_API_KEY_8", "VOYAGE_API_KEY_9"]:
        val = os.environ.get(var, "").strip()
        if val and val not in keys:
            keys.append(val)
    return keys

def embed_voyage_all(texts):
    """
    Embed ALL texts with TRUE round-robin key rotation.
    Each batch proactively uses the next key — no waiting for 429.
    429 fallback still handled if a key is exhausted mid-run.
    """
    import urllib.request, json as _json

    url = "https://api.voyageai.com/v1/embeddings"
    BATCH_SIZE = 16
    BACKOFF = [10, 15, 20, 25, 30, 40, 300, 50, 720, 90]
    INTER_BATCH_SLEEP = 25  # ~2.7 RPM per key — safely under 3 RPM limit

    api_keys = load_voyage_keys()
    if not api_keys:
        raise ValueError("No VOYAGE_API_KEY found in environment")

    log.info(f"  Voyage: {len(api_keys)} key(s) loaded — round-robin rotation every batch")

    all_embeddings = []
    total_batches = (len(texts) + BATCH_SIZE - 1) // BATCH_SIZE

    for batch_num, i in enumerate(range(0, len(texts), BATCH_SIZE)):
        batch = texts[i:i + BATCH_SIZE]

        # TRUE round-robin — rotate key on every batch proactively
        primary_key_idx = batch_num % len(api_keys)
        log.info(f"  Batch {batch_num+1}/{total_batches} ({len(batch)} chunks) [key {primary_key_idx+1}/{len(api_keys)}]")

        data = None
        last_err = None
        # Start rotation from the primary key for this batch
        current_key_idx = primary_key_idx

        for attempt, wait in enumerate(BACKOFF):
            active_key = api_keys[current_key_idx % len(api_keys)]
            headers = {
                "Content-Type": "application/json",
                "Authorization": f"Bearer {active_key}"
            }
            payload = _json.dumps({"input": batch, "model": "voyage-4"}).encode()
            req = urllib.request.Request(url, data=payload, headers=headers, method="POST")

            try:
                with urllib.request.urlopen(req) as resp:
                    data = _json.loads(resp.read())
                break  # success

            except urllib.error.HTTPError as e:
                if e.code == 429:
                    last_err = e
                    from_key = (current_key_idx % len(api_keys)) + 1
                    current_key_idx += 1
                    to_key = (current_key_idx % len(api_keys)) + 1
                    log.warning(f"  429 on key {from_key} (attempt {attempt+1}/{len(BACKOFF)}) — switching to key {to_key}, waiting {wait}s")
                    if attempt < len(BACKOFF) - 1:
                        time.sleep(wait)
                    else:
                        log.error("  All retries exhausted.")
                        raise
                else:
                    raise

            except urllib.error.URLError as e:
                last_err = e
                log.warning(f"  Network error (attempt {attempt+1}/{len(BACKOFF)}): {e.reason} — waiting {wait}s")
                if attempt < len(BACKOFF) - 1:
                    time.sleep(wait)
                else:
                    log.error("  Network error: all retries exhausted.")
                    raise

            except Exception as e:
                if "429" in str(e):
                    last_err = e
                    from_key = (current_key_idx % len(api_keys)) + 1
                    current_key_idx += 1
                    to_key = (current_key_idx % len(api_keys)) + 1
                    log.warning(f"  429 on key {from_key} (attempt {attempt+1}/{len(BACKOFF)}) — switching to key {to_key}, waiting {wait}s")
                    if attempt < len(BACKOFF) - 1:
                        time.sleep(wait)
                    else:
                        log.error("  All retries exhausted.")
                        raise
                else:
                    raise

        if data is None:
            raise last_err

        for item in sorted(data["data"], key=lambda x: x["index"]):
            all_embeddings.append(item["embedding"])

        if i + BATCH_SIZE < len(texts):
            time.sleep(INTER_BATCH_SLEEP)

    return np.array(all_embeddings, dtype=np.float32)

def embed_sentence(texts):
    from sentence_transformers import SentenceTransformer
    return SentenceTransformer("all-MiniLM-L6-v2").encode(texts, convert_to_numpy=True).astype(np.float32)

def db_paths(db_dir):
    d = Path(db_dir)
    d.mkdir(parents=True, exist_ok=True)
    return d / "index.faiss", d / "metadata.pkl"

def load_db(db_dir, embedding_backend="hash"):
    idx_path, meta_path = db_paths(db_dir)
    if idx_path.exists() and meta_path.exists():
        index = faiss.read_index(str(idx_path))
        with open(meta_path, "rb") as f:
            chunks = pickle.load(f)
        return index, chunks
    return faiss.IndexFlatIP(get_dim(embedding_backend)), []

def save_db(db_dir, index, chunks):
    idx_path, meta_path = db_paths(db_dir)
    faiss.write_index(index, str(idx_path))
    with open(meta_path, "wb") as f:
        pickle.dump(chunks, f)

def already_ingested(chunks, fhash):
    return any(c.file_hash == fhash for c in chunks)

def file_hash(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for block in iter(lambda: f.read(65536), b""): h.update(block)
    return h.hexdigest()[:16]

def ingest_file(path, index, chunks, cfg):
    fhash = file_hash(path)
    if already_ingested(chunks, fhash):
        log.info(f"  Skipping (already ingested): {path.name}")
        return 0

    log.info(f"  Ingesting: {path.name}  [plant_id={cfg.plant_id}]")
    sections = extract_document(path, cfg)
    if not sections:
        log.warning(f"  No text extracted: {path.name}")
        return 0

    new_chunks = []
    for section_text, section_label in sections:
        if not section_text.strip(): continue
        for raw in chunk_text(section_text, cfg.chunk_size, cfg.chunk_overlap, cfg.min_chunk_words):
            new_chunks.append(DocumentChunk(
                chunk_id=f"{fhash}_{len(new_chunks):04d}",
                text=raw,
                source_file=path.name,
                doc_type=path.suffix.lower().lstrip("."),
                page_or_sheet=section_label,
                chunk_index=len(new_chunks),
                total_chunks=0,
                file_hash=fhash,
                word_count=len(raw.split()),
                plant_id=cfg.plant_id))

    if not new_chunks: return 0
    for c in new_chunks: c.total_chunks = len(new_chunks)

    log.info(f"  {path.name} → {len(new_chunks)} chunks — embedding all at once...")

    all_texts = [c.text for c in new_chunks]

    if cfg.embedding_backend == "voyage":
        vecs = embed_voyage_all(all_texts).astype(np.float32)
    elif cfg.embedding_backend == "sentence":
        vecs = embed_sentence(all_texts).astype(np.float32)
    else:
        vecs = embed_hash(all_texts).astype(np.float32)

    faiss.normalize_L2(vecs)
    index.add(vecs)
    chunks.extend(new_chunks)
    log.info(f"  ✓ {path.name} → {len(new_chunks)} chunks added to DB")
    return len(new_chunks)

def ingest_directory(cfg):
    input_path = Path(cfg.input_dir)
    if not input_path.exists(): raise FileNotFoundError(f"Not found: {cfg.input_dir}")
    index, chunks = load_db(cfg.db_dir, cfg.embedding_backend)
    files = [f for f in input_path.rglob("*") if f.is_file() and f.suffix.lower() in cfg.supported_extensions]
    if not files:
        log.warning(f"No supported files found in {cfg.input_dir}")
        return {"files_found": 0, "chunks_added": 0}

    log.info(f"Found {len(files)} file(s)  [plant_id={cfg.plant_id}]")
    total, processed, skipped = 0, 0, 0

    for file_num, f in enumerate(files):
        log.info(f"--- File {file_num+1}/{len(files)}: {f.name} ---")
        added = ingest_file(f, index, chunks, cfg)
        if added > 0:
            total += added
            processed += 1
            save_db(cfg.db_dir, index, chunks)
            if file_num < len(files) - 1:
                log.info("  Sleeping 15s before next file...")
                time.sleep(15)
        else:
            skipped += 1

    log.info(f"Done. {processed} files processed, {skipped} skipped, {total} chunks added. DB total: {len(chunks)}")
    return {
        "files_found": len(files),
        "files_processed": processed,
        "files_skipped": skipped,
        "chunks_added": total,
        "total_chunks_in_db": len(chunks)
    }

def query_collection(question, cfg, top_k=5, plant_id=None):
    """
    plant_id: optional filter. If provided, only chunks tagged with this plant_id
    are eligible to be returned. If None, searches across all plants (default
    backward-compatible behavior — single-tenant demo still works unchanged).

    Implementation note: FAISS IndexFlatIP doesn't support native metadata
    filtering, so we over-fetch from the index and filter in Python. This is
    fine at hackathon/demo scale (hundreds-thousands of chunks); at real
    multi-tenant scale this would move to a vector DB with native filtering
    (e.g. Pinecone, Qdrant, Weaviate) using per-plant namespaces/collections.
    """
    index, chunks = load_db(cfg.db_dir, cfg.embedding_backend)
    if not chunks: return []
    BACKOFF = [10, 20, 30, 60, 120, 180]
    last_err = None
    for attempt, wait in enumerate(BACKOFF):
        try:
            if cfg.embedding_backend == "voyage":
                q_vec = embed_voyage_all([question]).astype(np.float32)
            elif cfg.embedding_backend == "sentence":
                q_vec = embed_sentence([question]).astype(np.float32)
            else:
                q_vec = embed_hash([question]).astype(np.float32)
            break
        except Exception as e:
            if "429" in str(e) and attempt < len(BACKOFF) - 1:
                log.warning(f"Rate limit on query (attempt {attempt+1}) — waiting {wait}s")
                time.sleep(wait)
                last_err = e
            else:
                raise
    else:
        raise last_err
    faiss.normalize_L2(q_vec)

    if plant_id is None:
        # No filter — original behavior, unchanged
        scores, indices = index.search(q_vec, min(top_k, len(chunks)))
        return [{"chunk": chunks[idx], "score": round(float(score)*100, 1)}
                for score, idx in zip(scores[0], indices[0]) if 0 <= idx < len(chunks)]

    # Filtered search: over-fetch then filter by plant_id, since FAISS
    # IndexFlatIP has no native metadata filter.
    fetch_k = min(len(chunks), max(top_k * 8, 50))
    scores, indices = index.search(q_vec, fetch_k)
    results = []
    for score, idx in zip(scores[0], indices[0]):
        if not (0 <= idx < len(chunks)): continue
        chunk = chunks[idx]
        # getattr fallback: old chunks ingested before plant_id existed
        # default to "default" so they remain queryable under that plant
        chunk_plant = getattr(chunk, "plant_id", "default")
        if chunk_plant == plant_id:
            results.append({"chunk": chunk, "score": round(float(score) * 100, 1)})
        if len(results) >= top_k:
            break
    return results

def watch_directory(cfg, poll_interval=10):
    log.info(f"Watching {cfg.input_dir} every {poll_interval}s. Ctrl+C to stop.")
    index, chunks = load_db(cfg.db_dir, cfg.embedding_backend)
    seen = {c.file_hash for c in chunks}
    while True:
        for f in Path(cfg.input_dir).rglob("*"):
            if not f.is_file() or f.suffix.lower() not in cfg.supported_extensions: continue
            fhash = file_hash(f)
            if fhash not in seen:
                added = ingest_file(f, index, chunks, cfg)
                if added > 0:
                    seen.add(fhash)
                    save_db(cfg.db_dir, index, chunks)
        time.sleep(poll_interval)

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="./docs")
    parser.add_argument("--db", default="./voyage_faiss_db")
    parser.add_argument("--embedding", default="voyage", choices=["hash","voyage","sentence"])
    parser.add_argument("--chunk-size", type=int, default=500)
    parser.add_argument("--overlap", type=int, default=50)
    parser.add_argument("--plant-id", default="default", help="Plant/site identifier for multi-tenant document tagging")
    parser.add_argument("--watch", action="store_true")
    parser.add_argument("--query", type=str)
    parser.add_argument("--top-k", type=int, default=5)
    args = parser.parse_args()
    cfg = IngestConfig(input_dir=args.input, db_dir=args.db, embedding_backend=args.embedding,
                       chunk_size=args.chunk_size, chunk_overlap=args.overlap, plant_id=args.plant_id)
    if args.query:
        for r in query_collection(args.query, cfg, args.top_k):
            c = r["chunk"]
            print(f"[{r['score']}%] {c.source_file} — {c.page_or_sheet}\n  {c.text[:200]}...")
    elif args.watch:
        watch_directory(cfg)
    else:
        print(json.dumps(ingest_directory(cfg), indent=2))

if __name__ == "__main__":
    main()